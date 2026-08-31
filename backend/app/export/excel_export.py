"""Excel export: Données / Seuils / Résultats / Comparatif tabs.

Only the raw measurements and the merged existant/projet widths are written as
values; every ratio and every comparatif delta/statut is a live Excel formula
(COUNTIFS/COUNT/IF) referencing the Seuils tab, so editing a threshold there
recalculates the whole workbook without reopening the tool.
"""

from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule

from app.models.domain import ComparisonRow, ElementType, Side, StateKind, Threshold, WidthSample

_HEADER_FILL = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
_INPUT_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
_AMELIORE_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
_DEGRADE_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_HEADER_FONT = Font(bold=True)

_SIDE_LABEL = {Side.GAUCHE: "Gauche", Side.DROITE: "Droite"}
_TYPE_LABEL = {
    ElementType.NON_UTILISE: "Non utilisé",
    ElementType.ACCOTEMENT: "Accotement",
    ElementType.TROTTOIR: "Trottoir",
    ElementType.BAU: "BAU",
    ElementType.CYCLE: "Cycle",
    ElementType.VOIE: "Voie",
    ElementType.TPC: "TPC",
}
_STATE_LABEL = {StateKind.EXISTANT: "Existant", StateKind.PROJET: "Projet"}


def _group_key(side: Side, element_type: ElementType, state: StateKind) -> str:
    return f"{_SIDE_LABEL[side]} - {_TYPE_LABEL[element_type]} - {_STATE_LABEL[state]}"


def build_workbook(
    samples: list[WidthSample],
    thresholds: list[Threshold],
    delta_seuil_m: float,
    comparison_rows: list[ComparisonRow] | None = None,
) -> Workbook:
    wb = Workbook()
    donnees_ws = wb.active
    donnees_ws.title = "Données"

    groups: dict[tuple[Side, ElementType, StateKind], dict[float, float]] = defaultdict(dict)
    for s in samples:
        if s.width_m is not None:
            groups[(s.side, s.element_type, s.state)][s.pk] = s.width_m

    master_pks = sorted({pk for g in groups.values() for pk in g})
    group_keys = sorted(groups.keys(), key=lambda k: (k[0].value, k[1].value, k[2].value))
    column_of_group = _write_donnees_sheet(donnees_ws, master_pks, group_keys, groups)

    seuils_ws = wb.create_sheet("Seuils")
    threshold_cell_by_type, delta_cell = _write_seuils_sheet(seuils_ws, thresholds, delta_seuil_m)

    resultats_ws = wb.create_sheet("Résultats")
    _write_resultats_sheet(resultats_ws, group_keys, column_of_group, threshold_cell_by_type, len(master_pks))

    if comparison_rows:
        comparatif_ws = wb.create_sheet("Comparatif")
        _write_comparatif_sheet(comparatif_ws, comparison_rows, delta_cell)

    return wb


def _write_donnees_sheet(ws: Worksheet, master_pks: list[float], group_keys, groups) -> dict:
    ws.cell(row=1, column=1, value="PK").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    column_of_group = {}
    for i, key in enumerate(group_keys):
        col = 2 + i
        column_of_group[key] = col
        cell = ws.cell(row=1, column=col, value=_group_key(*key))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = 22

    for row_idx, pk in enumerate(master_pks, start=2):
        ws.cell(row=row_idx, column=1, value=pk)
        for key in group_keys:
            width = groups[key].get(pk)
            if width is not None:
                ws.cell(row=row_idx, column=column_of_group[key], value=width)
    ws.freeze_panes = "B2"
    return column_of_group


def _write_seuils_sheet(ws: Worksheet, thresholds: list[Threshold], delta_seuil_m: float) -> tuple[dict, str]:
    ws.cell(row=1, column=1, value="Élément").font = _HEADER_FONT
    ws.cell(row=1, column=2, value="Réduit (m)").font = _HEADER_FONT
    ws.cell(row=1, column=3, value="Standard (m)").font = _HEADER_FONT
    threshold_by_type = {t.element_type: t for t in thresholds}
    row = 2
    cell_by_type: dict[ElementType, tuple[str, str]] = {}
    for element_type in ElementType:
        if element_type == ElementType.NON_UTILISE:
            continue
        t = threshold_by_type.get(element_type)
        ws.cell(row=row, column=1, value=_TYPE_LABEL[element_type])
        reduit_cell = ws.cell(row=row, column=2, value=t.reduit_m if t else None)
        standard_cell = ws.cell(row=row, column=3, value=t.standard_m if t else None)
        reduit_cell.fill = _INPUT_FILL
        standard_cell.fill = _INPUT_FILL
        cell_by_type[element_type] = (f"Seuils!$B${row}", f"Seuils!$C${row}")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Seuil de variation significative (m)")
    delta_input = ws.cell(row=row, column=2, value=delta_seuil_m)
    delta_input.fill = _INPUT_FILL
    delta_cell = f"Seuils!$B${row}"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    return cell_by_type, delta_cell


def _write_resultats_sheet(ws: Worksheet, group_keys, column_of_group, threshold_cell_by_type, n_pk_rows: int) -> None:
    headers = ["Côté", "Élément", "État", "N", "< Réduit %", "≥ Réduit < Standard %", "≥ Standard %"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    row = 2
    for key in group_keys:
        side, element_type, state = key
        if element_type not in threshold_cell_by_type or n_pk_rows == 0:
            continue
        col_letter = get_column_letter(column_of_group[key])
        data_range = f"Données!${col_letter}$2:${col_letter}${1 + n_pk_rows}"
        reduit_cell, standard_cell = threshold_cell_by_type[element_type]

        ws.cell(row=row, column=1, value=_SIDE_LABEL[side])
        ws.cell(row=row, column=2, value=_TYPE_LABEL[element_type])
        ws.cell(row=row, column=3, value=_STATE_LABEL[state])
        ws.cell(row=row, column=4, value=f"=COUNT({data_range})")
        ws.cell(
            row=row,
            column=5,
            value=f'=IFERROR(COUNTIFS({data_range},"<"&{reduit_cell})/COUNT({data_range})*100,0)',
        )
        ws.cell(
            row=row,
            column=6,
            value=(
                f'=IFERROR(COUNTIFS({data_range},">="&{reduit_cell},{data_range},"<"&{standard_cell})'
                f"/COUNT({data_range})*100,0)"
            ),
        )
        ws.cell(
            row=row,
            column=7,
            value=f'=IFERROR(COUNTIFS({data_range},">="&{standard_cell})/COUNT({data_range})*100,0)',
        )
        row += 1

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_comparatif_sheet(ws: Worksheet, comparison_rows: list[ComparisonRow], delta_cell: str) -> None:
    headers = ["PK", "Côté", "Élément", "Largeur existant (m)", "Largeur projet (m)", "Delta (m)", "Statut"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for i, row_data in enumerate(sorted(comparison_rows, key=lambda r: (r.side.value, r.element_type.value, r.pk)), start=2):
        ws.cell(row=i, column=1, value=row_data.pk)
        ws.cell(row=i, column=2, value=_SIDE_LABEL[row_data.side])
        ws.cell(row=i, column=3, value=_TYPE_LABEL[row_data.element_type])
        ws.cell(row=i, column=4, value=row_data.width_existant)
        ws.cell(row=i, column=5, value=row_data.width_projet)
        ws.cell(row=i, column=6, value=f'=IF(OR(D{i}="",E{i}=""),"",E{i}-D{i})')
        ws.cell(
            row=i,
            column=7,
            value=(
                f'=IF(F{i}="","",IF(F{i}>{delta_cell},"Amélioré",'
                f'IF(F{i}<-{delta_cell},"Dégradé","Inchangé")))'
            ),
        )

    last_row = 1 + len(comparison_rows)
    if last_row >= 2:
        status_range = f"$G$2:$G${last_row}"
        ws.conditional_formatting.add(
            status_range, FormulaRule(formula=['$G2="Amélioré"'], fill=_AMELIORE_FILL)
        )
        ws.conditional_formatting.add(
            status_range, FormulaRule(formula=['$G2="Dégradé"'], fill=_DEGRADE_FILL)
        )

    synth_col = 9
    ws.cell(row=1, column=synth_col, value="Synthèse").font = _HEADER_FONT
    labels_and_statuses = [("Amélioré", "Amélioré"), ("Inchangé", "Inchangé"), ("Dégradé", "Dégradé")]
    for i, (label, status) in enumerate(labels_and_statuses, start=2):
        ws.cell(row=i, column=synth_col, value=label)
        ws.cell(row=i, column=synth_col + 1, value=f'=COUNTIF($G$2:$G${max(last_row,2)},"{status}")')

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 20
